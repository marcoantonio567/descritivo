from openpyxl import load_workbook
from openpyxl.styles import Font ,PatternFill , Alignment
 



def editar_celula(celula, novo_valor,aba='quadro_resumo'):
    # Carregar o arquivo Excel
    arquivo = 'integracao.xlsx'
    workbook = load_workbook(arquivo)
    
    # Selecionar a aba desejada
    sheet = workbook[aba]
    
    # Editar a célula especificada
    sheet[celula] = novo_valor
    
    # Colocar o valor em negrito, com fonte 'Century Gothic' e tamanho 12
    custom_font = Font(name='Century Gothic', size=12, bold=True)
    sheet[celula].font = custom_font
    
    # Salvar o arquivo
    workbook.save(arquivo)
    workbook.close()
    print(f"Célula {celula} editada com sucesso!")
def contar_dados_vazios(aba='propietarios',coluna='C'):
    caminho_arquivo = 'integracao.xlsx'
    # Carrega a planilha Excel
    workbook = load_workbook(caminho_arquivo)
    # Selecione a aba ativa ou a aba de interesse
    sheet = workbook[aba]

    # Verifica as células do intervalo C3 a C12
    celulas_vazias = []
    for linha in range(3, 13):
        celula = sheet[f"{coluna}{linha}"]
        if celula.value is None:
            celulas_vazias.append(f"{coluna}{linha}")

    # Retorna o número de células vazias
    workbook.close()
    return len(celulas_vazias)
def numero_Da_celula_preenchimento(aba='quadro_resumo'):
    arquivo_excel = 'integracao.xlsx'
    # Carregar a planilha do Excel
    workbook = load_workbook(arquivo_excel)
    # Selecionar a aba desejada (modifique para o nome da sua aba se necessário)
    aba = workbook[aba]

    # Iterar pelas linhas da coluna A para encontrar a última célula preenchida
    ultima_linha = 0
    for linha in range(1, aba.max_row + 1):
        if aba[f'A{linha}'].value is not None:
            ultima_linha = linha
    workbook.close()
    return ultima_linha +1
def mesclar_celulas(linha, intervalos, aba='quadro_resumo', centralizar=False):
    nome_arquivo = 'integracao.xlsx'
    # Carregar o arquivo Excel existente
    workbook = load_workbook(nome_arquivo)

    # Selecionar a planilha específica
    sheet = workbook[aba]

    # Definir o preenchimento branco para as células mescladas
    preenchimento_branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Iterar sobre os intervalos para mesclar as células
    for inicio, fim in intervalos:
        # Criar o intervalo de mesclagem usando a linha especificada
        intervalo = f"{inicio}{linha}:{fim}{linha}"
        sheet.merge_cells(intervalo)

        # Aplicar a cor de fundo branca à célula "superior esquerda"
        celula_superior_esquerda = sheet[f"{inicio}{linha}"]
        celula_superior_esquerda.fill = preenchimento_branco

        # Se o parâmetro centralizar for True, aplicar o alinhamento centralizado
        if centralizar:
            alinhamento_centralizado = Alignment(horizontal='center', vertical='center')
            celula_superior_esquerda.alignment = alinhamento_centralizado

    # Salvar as alterações no arquivo Excel
    workbook.save(nome_arquivo)
    workbook.close()
def ler_propietarios(quantidade):
    linha_maxima =  quantidade +2
    caminho_arquivo = 'integracao.xlsx'
    nome_planilha = 'propietarios'
    # Abre o arquivo Excel com data_only=True para ler o valor das fórmulas
    workbook = load_workbook(caminho_arquivo, data_only=True)
    # Seleciona a pagina especificada
    pagina = workbook[nome_planilha]
    # Obtém o valor da célula especificada
    propietarios = []
    for linha in pagina.iter_rows(values_only=True,min_col=3,min_row=3,max_row=linha_maxima):
        nome = linha[0]
        cpf_cnpj = linha[1]
        propietarios.append((nome,cpf_cnpj))
    workbook.close()
    return propietarios
def ler_matriculas(quantidade):
    linha_maxima =  quantidade +2
    caminho_arquivo = 'integracao.xlsx'
    nome_planilha = 'propietarios'
    # Abre o arquivo Excel com data_only=True para ler o valor das fórmulas
    workbook = load_workbook(caminho_arquivo, data_only=True)
    # Seleciona a pagina especificada
    pagina = workbook[nome_planilha]
    # Obtém o valor da célula especificada
    propietarios = []
    for linha in pagina.iter_rows(values_only=True,min_col=7,min_row=3,max_row=linha_maxima):
        matricula = linha[0]
        identificacao = linha[1]
        area_ha = linha[2]
        area_contruida = linha[3]
        valor_mercado = linha[4]
        liquidacao_forcada = linha[5]
        propietarios.append((matricula,identificacao,area_ha,area_contruida,valor_mercado,liquidacao_forcada))
    return propietarios
def fazer_quadro_resumo():
    quantidade_propontes = 10 - contar_dados_vazios() # resposta 3
    lista_proprietarios = ler_propietarios(quantidade_propontes)
    letra_incio = 'A'
    letra_meio = 'C'
    letra_fim = 'J'
    palavras = ['Agencia:','Identificação:','Município:']
    intervalos_para_mesclar = [('A', 'B'), ('C', 'I'),('J','N')]
    
    for item in lista_proprietarios:
        numero_ultima_celula = numero_Da_celula_preenchimento()
        ultima_celula = letra_incio + str(numero_ultima_celula)
        editar_celula(ultima_celula,'Proprietário:')
        editar_celula(letra_meio + str(numero_ultima_celula),str(item[0]))
        editar_celula(letra_fim + str(numero_ultima_celula),'CPF: '+str(item[1]))
        mesclar_celulas(numero_ultima_celula,intervalos_para_mesclar)
        
    for palavra in palavras:
        numero_ultima_celula = numero_Da_celula_preenchimento()
        ultima_celula = letra_incio + str(numero_ultima_celula)
        editar_celula(ultima_celula,palavra)
        mesclar_celulas(numero_ultima_celula,intervalos_para_mesclar)

    editar_celula(letra_fim+str(numero_ultima_celula),'UF:')
def fazer_imovel():
    letra_incio = 'A'
    letra_meio_1 = 'D'
    letra_meio_2 = 'G'
    letra_fim = 'J'

    quantidade_matriculas = 10 - contar_dados_vazios(coluna='G')
    dados_imoveis = ler_matriculas(quantidade_matriculas)
    intervalos_para_mesclar = [('A', 'C'), ('D', 'F'),('G','I'),('J','M')]

    for item in dados_imoveis:
        numero_ultima_celula = numero_Da_celula_preenchimento(aba='imovel')
        ultima_celula = letra_incio + str(numero_ultima_celula)
        editar_celula(ultima_celula,str(item[0]),aba='imovel')
        editar_celula(letra_meio_1+str(numero_ultima_celula),str(item[1]),aba='imovel')
        editar_celula(letra_meio_2+str(numero_ultima_celula),str(item[2]),aba='imovel')
        editar_celula(letra_fim+str(numero_ultima_celula),str(item[3]),aba='imovel')
        mesclar_celulas(numero_ultima_celula,intervalos_para_mesclar,aba='imovel',centralizar=True)
def fazer_valores():
    letra_incio = 'A'
    letra_meio = 'F'
    letra_fim = 'K'

    quantidade_matriculas = 10 - contar_dados_vazios(coluna='G')
    dados_imoveis = ler_matriculas(quantidade_matriculas)
    intervalos_para_mesclar = [('A', 'C'), ('F', 'H'),('K','M')]
    
    for item in dados_imoveis:
        numero_ultima_celula = numero_Da_celula_preenchimento(aba='valores')
        ultima_celula = letra_incio + str(numero_ultima_celula)
        editar_celula(ultima_celula,str(item[0]),aba='valores')
        editar_celula(letra_meio+str(numero_ultima_celula),str(item[4]),aba='valores')
        editar_celula(letra_fim+str(numero_ultima_celula),str(item[5]),aba='valores')
        mesclar_celulas(numero_ultima_celula,intervalos_para_mesclar,aba='valores',centralizar=True)

fazer_valores()
fazer_imovel()
fazer_quadro_resumo()
