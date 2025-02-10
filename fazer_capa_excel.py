from openpyxl import load_workbook



arquivo = 'integracao.xlsx'
workbook = load_workbook(arquivo,data_only=True)

def contar_dados_vazios(aba='propietarios',coluna='C'):
    # Selecione a aba ativa ou a aba de interesse
    sheet = workbook[aba]

    # Verifica as células do intervalo C3 a C12
    celulas_vazias = []
    for linha in range(3, 13):
        celula = sheet[f"{coluna}{linha}"]
        if celula.value is None:
            celulas_vazias.append(f"{coluna}{linha}")

    
    return len(celulas_vazias)
def ler_propietarios(quantidade):
    linha_maxima =  quantidade +2
    
    nome_planilha = 'propietarios'

    # Seleciona a pagina especificada
    pagina = workbook[nome_planilha]
    # Obtém o valor da célula especificada
    propietarios = []
    for linha in pagina.iter_rows(values_only=True,min_col=1,min_row=3,max_row=linha_maxima):
        genero = linha[0]
        porcentagem = linha[1]
        nome = linha[2]
        cpf_cnpj = linha[3]
        casado = linha[4]
        propietarios.append((nome,cpf_cnpj,porcentagem,genero,casado))
   
    return propietarios
def ler_matriculas(quantidade):
    linha_maxima =  quantidade +2
    
    nome_planilha = 'propietarios'
    
    # Seleciona a pagina especificada
    pagina = workbook[nome_planilha]
    # Obtém o valor da célula especificada
    propietarios = []
    for linha in pagina.iter_rows(values_only=True,min_col=7,min_row=3,max_row=linha_maxima):
        matricula = linha[0]
        identificacao = linha[1]
        area_ha = linha[2]
        area_contruida = linha[3]
        valor_mercado = linha[4]
        liquidacao_forcada = linha[5]
        nome_fazendas = linha[6]
        data_emissao = linha[7]
        car = linha[8]
        propietarios.append((matricula,identificacao,area_ha,area_contruida,valor_mercado,liquidacao_forcada,nome_fazendas,data_emissao,car))
    return propietarios
def ler_casamento(quantidade):
    linha_maxima = quantidade+14
    nome_planilha = 'propietarios'

    # Seleciona a pagina especificada
    pagina = workbook[nome_planilha]
    # Obtém o valor da célula especificada
    casamentos = []
    for linha in pagina.iter_rows(values_only=True,min_col=2,min_row=15,max_row=linha_maxima):
        genero = linha[0]
        nome = linha[1]
        cpf = linha[2]
        casamentos.append((genero,nome,cpf))
    return casamentos

    
workbook.save(arquivo)
workbook.close()

